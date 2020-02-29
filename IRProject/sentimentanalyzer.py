# -*- coding: utf-8 -*-
"""
Created on Mon Jan 13 20:23:23 2020

@author: mikeygh
"""

import pandas as pd
import os
from nltk.tokenize import TweetTokenizer
import nltk
import re

nltk.download('stopwords')
nltk.download('words')
from nltk.corpus import stopwords
from openpyxl import load_workbook
from nltk.stem import PorterStemmer
#from spellchecker import SpellChecker
import xlsxwriter
import os.path
from os import path

english_vocab = set(w.lower() for w in nltk.corpus.words.words())

#spell = SpellChecker()
sentimental_dictionary = {}
tknzr = TweetTokenizer(preserve_case=True, reduce_len=True, strip_handles=True)
ps = PorterStemmer()
cachedStopWords = stopwords.words("english")

row = 0
column = 0
createdGlossary = False
if (path.exists("Sentimental Glossary.xlsx")):
    createdGlossary = True
else:
    workbook_glossary = xlsxwriter.Workbook('Sentimental Glossary.xlsx')
    worksheet_glossary = workbook_glossary.add_worksheet()

dictionary = {}
vocabulary = dict()


def iter_rows(ws):
    result = []
    for row in ws.iter_rows():
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        result.append(rowlist)
    return result


def word_count(str):
    for word in str:
        if word in vocabulary:
            vocabulary[word] += 1
        else:
            vocabulary[word] = 1


workbook = load_workbook(filename='../data.xlsx', read_only=True)
worksheet = workbook['sheet']
workbook1 = load_workbook(filename='../SentimentalGlossary-Copy.xlsx', read_only=True)
worksheet1 = workbook1['Sheet1']

fileList = (list(iter_rows(worksheet)))
fileList1 = (list(iter_rows(worksheet1)))
geoLocation = []
tweet = []

for col in fileList:
    geoLocation.append(col[5])  # 1 is column index
    tweet.append(col[6])  # 2 is column index
for col in fileList1:
    sentimental_dictionary[col[0]] = col[1]

counter = 2
tweetgrade = {}
for i in tweet:
    currentTweet = tknzr.tokenize(i.casefold())
    currentTweet = [word for word in currentTweet if word not in cachedStopWords]
    currentTweet = [word for word in currentTweet if word in english_vocab]
    currentTweet = [s.strip('.') for s in currentTweet]
    currentTweet = [s.replace('.', '') for s in currentTweet]
    currentTweet = [s.strip('#') for s in currentTweet]
    currentTweet = [s.replace('#', '') for s in currentTweet]
    currentTweet = [s.strip(':') for s in currentTweet]
    currentTweet = [s.replace(':', '') for s in currentTweet]
    currentTweet = [s.strip('!') for s in currentTweet]
    currentTweet = [s.replace('!', '') for s in currentTweet]
    currentTweet = [s.strip('?') for s in currentTweet]
    currentTweet = [s.replace('?', '') for s in currentTweet]
    currentTweet = [s.strip('\\') for s in currentTweet]
    currentTweet = [s.replace('\\', '') for s in currentTweet]
    currentTweet = [s.strip('/') for s in currentTweet]
    currentTweet = [s.replace('/', '') for s in currentTweet]
    currentTweet = [s.strip('(') for s in currentTweet]
    currentTweet = [s.replace('(', '') for s in currentTweet]
    currentTweet = [s.strip(')') for s in currentTweet]
    currentTweet = [s.replace(')', '') for s in currentTweet]
    currentTweet = [s.strip('*') for s in currentTweet]
    currentTweet = [s.replace('*', '') for s in currentTweet]
    currentTweet = [s.strip(',') for s in currentTweet]
    currentTweet = [s.replace(',', '') for s in currentTweet]
    currentTweet = [s.strip('"') for s in currentTweet]
    currentTweet = [s.replace('"', '') for s in currentTweet]
    currentTweet = [s.strip('-') for s in currentTweet]
    currentTweet = [s.replace('-', '') for s in currentTweet]
    currentTweet = [s.strip('+') for s in currentTweet]
    currentTweet = [s.replace('+', '') for s in currentTweet]
    currentTweet = [s.strip('&') for s in currentTweet]
    currentTweet = [s.replace('&', '') for s in currentTweet]
    currentTweet = [s.strip('$') for s in currentTweet]
    currentTweet = [s.replace('$', '') for s in currentTweet]
    currentTweet = [s.strip(';') for s in currentTweet]
    currentTweet = [s.replace(';', '') for s in currentTweet]
    pattern = '[0-9]'
    currentTweet = [re.sub(pattern, '', i) for i in currentTweet]
    currentTweet = filter(lambda i: len(i) > 2, currentTweet)
    currentTweet = list(filter(None, currentTweet))
   # currentTweet = [spell.correction(i) for i in currentTweet]
    currentTweet = [ps.stem(i) for i in currentTweet]
    for word in currentTweet:
        if word in sentimental_dictionary.keys():
            tweetgrade[counter]=tweetgrade[counter]+sentimental_dictionary.get(word)
    counter=counter+1
print(tweetgrade)

