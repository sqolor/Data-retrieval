# -*- coding: utf-8 -*-
"""
Created on Mon Jan 13 20:23:23 2020

@author: mikeyg
"""

import pandas as pd
import os

from nltk.tokenize import TweetTokenizer
import nltk
import re

from wordcloud import WordCloud, STOPWORDS

nltk.download('stopwords')
from nltk.corpus import stopwords
from openpyxl import load_workbook
from nltk.stem import PorterStemmer
import numpy as np
import pandas as pd
from os import path
from PIL import Image
import wordcloud

tknzr = TweetTokenizer(preserve_case=True, reduce_len=True, strip_handles=True)
ps = PorterStemmer()
cachedStopWords = stopwords.words("english")

dictionary = {}


def iter_rows(ws):
    result = []
    for row in ws.iter_rows():
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        result.append(rowlist)
    return result


workbook = load_workbook(filename='../data.xlsx', read_only=True)
worksheet = workbook['sheet']

fileList = (list(iter_rows(worksheet)))
geoLocation = []
tweet = []
j = 2
text = ""
text1 = ""
text2 = ""
text3 = ""
tweetmap = dict()
geomap = dict()
for col in fileList:
    geoLocation.append(col[5])  # 1 is column index
    tweet.append(col[6])  # 2 is column index
    tweetmap[j] = col[6]
    geomap[j] = col[5]
    j = j + 1
    text += " " + col[6]
    if "Central Time (US & Canada)" in col[5]: text1 += " " + col[6]
    if "Eastern Time (US & Canada)" in col[5]: text2 += " " + col[6]
    if "Pacific Time (US & Canada)" in col[5]: text3 += " " + col[6]
counter = 2
# text = tweetmap[23]
stopwords = set(STOPWORDS)
stopwords.update(["https", "http", "ud83d", "ude2d", "ude02", "co"])

# Create and generate a word cloud image:
wordcloud = WordCloud(stopwords=stopwords, max_font_size=50, max_words=100, background_color="white").generate(text)
wordcloud.to_file("first_review.png")
wordcloud1 = WordCloud(stopwords=stopwords, max_font_size=50, max_words=100, background_color="white").generate(text1)
wordcloud1.to_file("second_review.png")
wordcloud2 = WordCloud(stopwords=stopwords, max_font_size=50, max_words=100, background_color="white").generate(text2)
wordcloud2.to_file("third_review.png")
wordcloud3 = WordCloud(stopwords=stopwords, max_font_size=50, max_words=100, background_color="white").generate(text3)
wordcloud3.to_file("fourth_review.png")


