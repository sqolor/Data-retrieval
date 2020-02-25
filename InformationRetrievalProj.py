# -*- coding: utf-8 -*-
"""
Created on Mon Jan 13 20:23:23 2020

@author: benic
"""

import pandas as pd
import os
from nltk.tokenize import TweetTokenizer
import nltk
import re
nltk.download('stopwords')
nltk.download('words')
from nltk.corpus import stopwords
from openpyxl import load_workbook
from nltk.stem import PorterStemmer
from spellchecker import SpellChecker
import xlsxwriter 
import os.path
from os import path

english_vocab = set(w.lower() for w in nltk.corpus.words.words())

spell = SpellChecker()
sentimental_dictionary={}
tknzr = TweetTokenizer(preserve_case=True, reduce_len=True, strip_handles=True)
ps = PorterStemmer()
cachedStopWords = stopwords.words("english")
import csv



row = 0
column = 0
bool createdGlossary = False
if(path.exists("Sentimental Glossary.xlsx")):
    createdGlossary = True
else:
    workbook_glossary = xlsxwriter.Workbook('Sentimental Glossary.xlsx') 
    worksheet_glossary = workbook_glossary.add_worksheet() 

dictionary = {}
vocabulary = dict()
def iter_rows(ws):
    result=[]
    for row in ws.iter_rows():
        rowlist = []
        for cell in row:
            rowlist.append(cell.value)
        result.append(rowlist)
    return result

def word_count(str):

    for word in str:
        if word in vocabulary:
            vocabulary[word] += 1
        else:
            vocabulary[word] = 1
workbook = load_workbook(filename='data.xlsx', read_only = True)
worksheet = workbook['sheet']

fileList =  (list(iter_rows(worksheet)))
geoLocation = []
tweet = []

for col in fileList:
    geoLocation.append(col[5])#1 is column index
    tweet.append(col[6])#2 is column index

counter = 2
for i in tweet:
    currentTweet = tknzr.tokenize(i.casefold())
    currentTweet = [word for word in currentTweet if word not in cachedStopWords]
    currentTweet = [word for word in currentTweet if word in english_vocab]
    currentTweet = [s.strip('.') for s in currentTweet] 
    currentTweet = [s.replace('.', '') for s in currentTweet]
    currentTweet = [s.strip('#') for s in currentTweet] 
    currentTweet = [s.replace('#', '') for s in currentTweet]
    currentTweet = [s.strip(':') for s in currentTweet] 
    currentTweet = [s.replace(':', '') for s in currentTweet]
    currentTweet = [s.strip('!') for s in currentTweet] 
    currentTweet = [s.replace('!', '') for s in currentTweet]
    currentTweet = [s.strip('?') for s in currentTweet] 
    currentTweet = [s.replace('?', '') for s in currentTweet]
    currentTweet = [s.strip('\\') for s in currentTweet] 
    currentTweet = [s.replace('\\', '') for s in currentTweet]
    currentTweet = [s.strip('/') for s in currentTweet] 
    currentTweet = [s.replace('/', '') for s in currentTweet]
    currentTweet = [s.strip('(') for s in currentTweet] 
    currentTweet = [s.replace('(', '') for s in currentTweet]
    currentTweet = [s.strip(')') for s in currentTweet] 
    currentTweet = [s.replace(')', '') for s in currentTweet]
    currentTweet = [s.strip('*') for s in currentTweet] 
    currentTweet = [s.replace('*', '') for s in currentTweet]
    currentTweet = [s.strip(',') for s in currentTweet] 
    currentTweet = [s.replace(',', '') for s in currentTweet]
    currentTweet = [s.strip('"') for s in currentTweet] 
    currentTweet = [s.replace('"', '') for s in currentTweet]
    currentTweet = [s.strip('-') for s in currentTweet] 
    currentTweet = [s.replace('-', '') for s in currentTweet]
    currentTweet = [s.strip('+') for s in currentTweet] 
    currentTweet = [s.replace('+', '') for s in currentTweet]
    currentTweet = [s.strip('&') for s in currentTweet] 
    currentTweet = [s.replace('&', '') for s in currentTweet]
    currentTweet = [s.strip('$') for s in currentTweet] 
    currentTweet = [s.replace('$', '') for s in currentTweet]
    currentTweet = [s.strip(';') for s in currentTweet] 
    currentTweet = [s.replace(';', '') for s in currentTweet]
    pattern = '[0-9]'
    currentTweet = [re.sub(pattern, '', i) for i in currentTweet]
    currentTweet = filter (lambda i: len (i) > 2, currentTweet)
    currentTweet = list(filter(None, currentTweet))
    currentTweet = [spell.correction(i) for i in currentTweet]
    currentTweet = [ps.stem(i) for i in currentTweet]
    word_count(currentTweet)
    print(currentTweet)


print(vocabulary)
if(createdGlossary is False):
    for key in vocabulary.keys():
    worksheet_glossary.write(row, column, key) 
    row+=1

    

workbook_glossary.close() 
